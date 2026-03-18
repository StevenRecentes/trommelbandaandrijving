"""
Import Ammdrive workbook data into db_tba catalog tables.

Usage:
  python scripts/import_ammdrive.py --dry-run
  python scripts/import_ammdrive.py --apply
  python scripts/import_ammdrive.py --apply --file "Project_uiteenzetting/Ammdrive overzicht en berekening.xlsx"
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SQL_DIR = PROJECT_ROOT / "sql"
if str(SQL_DIR) not in sys.path:
    sys.path.insert(0, str(SQL_DIR))

from database import get_db


DEFAULT_XLSX_PATH = Path("Project_uiteenzetting") / "Ammdrive overzicht en berekening.xlsx"
RULMECA_CANONICAL = "Rulmeca"


def canonical_supplier_name(value: Any) -> str:
    name = str(value or "").strip()
    lowered = name.lower().replace(" ", "")
    rulmeca_variants = {
        "rulmeca",
        "rulmecaspa",
        "rulmecas.p.a.",
        "rulmecas.p.a",
        "rulmecas.p.a",
    }
    if lowered in rulmeca_variants:
        return RULMECA_CANONICAL
    return name or "Onbekend"


def as_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text or text.lower() in {"#nb", "#n/b", "niet beschikbaar", "none"}:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def as_int(value: Any) -> int | None:
    numeric = as_number(value)
    if numeric is None:
        return None
    return int(round(numeric))


def oil_type_for_code(code: str) -> str:
    upper = str(code or "").upper()
    if "DER" in upper:
        return "olievrij"
    return "oliegevuld"


@dataclass
class ParsedData:
    leveranciers: list[str] = field(default_factory=list)
    uitvoering_types: list[dict[str, Any]] = field(default_factory=list)
    Aansluiting_types: list[dict[str, Any]] = field(default_factory=list)
    band_types: list[dict[str, Any]] = field(default_factory=list)
    motortypes: list[dict[str, Any]] = field(default_factory=list)
    motor_specs: list[dict[str, Any]] = field(default_factory=list)
    compatibiliteit: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def parse_workbook(workbook_path: Path) -> ParsedData:
    wb = load_workbook(workbook_path, data_only=True)
    parsed = ParsedData()

    parsed.uitvoering_types = [
        {"naam": "RVS mantel met groeven", "code": "rvs_mantel"},
        {"naam": "RVS mantel met spie en sprockets", "code": "sprocket"},
        {"naam": "RVS mantel met gegroefde bekleding", "code": "bekleding"},
    ]

    parsed.Aansluiting_types = _parse_Aansluitingen(wb["Uitv_tm"])
    parsed.band_types, band_rows = _parse_bandtypes(wb["Bandgegevens"])
    specs = _parse_motor_specs(wb["tech_specs TM"], wb["Blad1"])

    supplier_names = sorted({canonical_supplier_name(item["leverancier"]) for item in specs})
    parsed.leveranciers = supplier_names

    type_accumulator: dict[tuple[str, str], dict[str, Any]] = {}
    for item in specs:
        supplier = canonical_supplier_name(item["leverancier"])
        code = str(item["motortype_code"]).strip()
        key = (supplier, code)
        diameter = item["diameter_nominaal_mm"]
        lengte = item["lengte_mm"]
        current = type_accumulator.get(key)
        if not current:
            type_accumulator[key] = {
                "leverancier": supplier,
                "code": code,
                "diameter_nominaal_mm": diameter,
                "lengte_min_mm": lengte,
                "lengte_max_mm": lengte,
            }
            continue
        if diameter is not None:
            current["diameter_nominaal_mm"] = diameter
        if lengte is not None:
            if current["lengte_min_mm"] is None or lengte < current["lengte_min_mm"]:
                current["lengte_min_mm"] = lengte
            if current["lengte_max_mm"] is None or lengte > current["lengte_max_mm"]:
                current["lengte_max_mm"] = lengte

    parsed.motortypes = sorted(
        type_accumulator.values(),
        key=lambda row: (row["leverancier"], row["code"]),
    )

    parsed.motor_specs = [
        {
            "leverancier": canonical_supplier_name(item["leverancier"]),
            "motortype_code": str(item["motortype_code"]).strip(),
            "vermogen_w": item["vermogen_w"],
            "snelheid_ms": item["snelheid_ms"],
            "polen": item["polen"],
            "force_n": item["force_n"],
            "torque_nm": item["torque_nm"],
            "olie_type": oil_type_for_code(str(item["motortype_code"]).strip()),
        }
        for item in specs
    ]

    motortype_diameters = [
        {
            "leverancier": row["leverancier"],
            "code": row["code"],
            "diameter_nominaal_mm": row["diameter_nominaal_mm"],
        }
        for row in parsed.motortypes
    ]
    parsed.compatibiliteit.extend(_build_compatibiliteit(band_rows, motortype_diameters, parsed.warnings))
    return parsed


def _parse_Aansluitingen(ws) -> list[dict[str, Any]]:
    rows = []
    seen_codes: set[str] = set()
    for row_idx in range(4, ws.max_row + 1):
        raw = ws.cell(row_idx, 2).value
        text = str(raw or "").strip()
        if not text:
            continue
        if text[0].upper() == "T" and " " in text:
            code = text.split(" ", 1)[0].strip()
            if code in seen_codes:
                continue
            seen_codes.add(code)
            rows.append({"naam": text, "code": code})
    return rows


def _parse_bandtypes(ws) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    band_types: list[dict[str, Any]] = []
    band_rows: list[dict[str, Any]] = []
    for row_idx in range(5, 13):
        naam = str(ws.cell(row_idx, 3).value or "").strip()
        if not naam:
            continue
        band = {
            "naam": naam,
            "steek_mm": as_number(ws.cell(row_idx, 4).value),
            "dikte_tand_mm": as_number(ws.cell(row_idx, 5).value),
            "dikte_band_mm": as_number(ws.cell(row_idx, 6).value),
            "wrijvingscoeff_rvs_droog": as_number(ws.cell(row_idx, 7).value),
            "wrijvingscoeff_rvs_nat": as_number(ws.cell(row_idx, 8).value),
        }
        band_types.append(band)
        band_rows.append(
            {
                "naam": naam,
                "diameter_blocks": [
                    {
                        "diameter": as_number(ws.cell(row_idx, 12).value),
                        "rvs_tanden": as_int(ws.cell(row_idx, 13).value),
                        "sprocket_tanden": as_int(ws.cell(row_idx, 14).value),
                        "bekleding_tanden": as_int(ws.cell(row_idx, 15).value),
                        "rvs_pcd": as_number(ws.cell(row_idx, 16).value),
                        "sprocket_pcd": as_number(ws.cell(row_idx, 17).value),
                        "bekleding_pcd": as_number(ws.cell(row_idx, 18).value),
                    },
                    {
                        "diameter": as_number(ws.cell(row_idx, 19).value),
                        "rvs_tanden": as_int(ws.cell(row_idx, 20).value),
                        "sprocket_tanden": as_int(ws.cell(row_idx, 21).value),
                        "bekleding_tanden": as_int(ws.cell(row_idx, 22).value),
                        "rvs_pcd": as_number(ws.cell(row_idx, 23).value),
                        "sprocket_pcd": as_number(ws.cell(row_idx, 24).value),
                        "bekleding_pcd": as_number(ws.cell(row_idx, 25).value),
                    },
                    {
                        "diameter": as_number(ws.cell(row_idx, 26).value),
                        "rvs_tanden": as_int(ws.cell(row_idx, 27).value),
                        "sprocket_tanden": as_int(ws.cell(row_idx, 28).value),
                        "bekleding_tanden": as_int(ws.cell(row_idx, 29).value),
                        "rvs_pcd": as_number(ws.cell(row_idx, 30).value),
                        "sprocket_pcd": as_number(ws.cell(row_idx, 31).value),
                        "bekleding_pcd": as_number(ws.cell(row_idx, 32).value),
                    },
                ],
            }
        )
    return band_types, band_rows


def _parse_motor_specs(tech_ws, blad_ws) -> list[dict[str, Any]]:
    specs: dict[tuple[Any, ...], dict[str, Any]] = {}

    def add_row(
        leverancier: Any,
        motortype_code: Any,
        diameter: Any,
        vermogen_w: Any,
        polen: Any,
        lengte_mm: Any,
        snelheid_ms: Any,
        force_n: Any = None,
        torque_nm: Any = None,
    ) -> None:
        supplier = canonical_supplier_name(leverancier)
        code = str(motortype_code or "").strip()
        if not supplier or not code:
            return
        diameter_num = as_number(diameter)
        power_num = as_number(vermogen_w)
        speed_num = as_number(snelheid_ms)
        if power_num is None or speed_num is None:
            return
        poles_num = as_int(polen)
        length_num = as_number(lengte_mm)
        force_num = as_number(force_n)
        if force_num is None and speed_num:
            force_num = power_num / speed_num
        torque_num = as_number(torque_nm)
        if torque_num is None and force_num is not None and diameter_num:
            torque_num = (diameter_num / 2.0) * (force_num / 1000.0)
        key = (
            supplier,
            code,
            round(power_num, 6),
            round(speed_num, 6),
            poles_num or 0,
            round(length_num, 3) if length_num is not None else None,
        )
        specs[key] = {
            "leverancier": supplier,
            "motortype_code": code,
            "diameter_nominaal_mm": diameter_num,
            "vermogen_w": round(power_num, 6),
            "snelheid_ms": round(speed_num, 6),
            "polen": poles_num,
            "lengte_mm": round(length_num, 3) if length_num is not None else None,
            "force_n": round(force_num, 6) if force_num is not None else None,
            "torque_nm": round(torque_num, 6) if torque_num is not None else None,
        }

    for row_idx in range(4, tech_ws.max_row + 1):
        add_row(
            leverancier=tech_ws.cell(row_idx, 2).value,
            motortype_code=tech_ws.cell(row_idx, 1).value,
            diameter=tech_ws.cell(row_idx, 3).value,
            vermogen_w=tech_ws.cell(row_idx, 4).value,
            polen=tech_ws.cell(row_idx, 5).value,
            lengte_mm=tech_ws.cell(row_idx, 10).value,
            snelheid_ms=tech_ws.cell(row_idx, 6).value,
            force_n=tech_ws.cell(row_idx, 7).value,
            torque_nm=tech_ws.cell(row_idx, 8).value,
        )

    for row_idx in range(1, blad_ws.max_row + 1):
        add_row(
            leverancier=blad_ws.cell(row_idx, 1).value,
            motortype_code=blad_ws.cell(row_idx, 2).value,
            diameter=blad_ws.cell(row_idx, 3).value,
            vermogen_w=blad_ws.cell(row_idx, 4).value,
            polen=blad_ws.cell(row_idx, 5).value,
            lengte_mm=blad_ws.cell(row_idx, 6).value,
            snelheid_ms=blad_ws.cell(row_idx, 7).value,
        )

    return list(specs.values())


def _build_compatibiliteit(
    band_rows: list[dict[str, Any]],
    motortype_diameters: list[dict[str, Any]],
    warnings: list[str],
) -> list[dict[str, Any]]:
    compat: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()

    for band in band_rows:
        for motortype in motortype_diameters:
            diameter = motortype["diameter_nominaal_mm"]
            if diameter is None:
                continue
            blocks = [b for b in band["diameter_blocks"] if b["diameter"] is not None]
            if not blocks:
                continue
            nearest = min(blocks, key=lambda b: abs(float(diameter) - float(b["diameter"])))
            delta = abs(float(diameter) - float(nearest["diameter"]))
            if 12 < delta <= 30:
                warnings.append(
                    f"Bandcompat benaderd voor {motortype['code']} ({diameter} mm) via dichtstbijzijnde PCD-matrix {nearest['diameter']} mm"
                )
            if delta > 30:
                warnings.append(
                    f"Geen directe bandcompat voor {motortype['code']} ({diameter} mm) op band '{band['naam']}'"
                )
                continue

            for uitvoering, tanden_key, pcd_key in [
                ("rvs_mantel", "rvs_tanden", "rvs_pcd"),
                ("sprocket", "sprocket_tanden", "sprocket_pcd"),
                ("bekleding", "bekleding_tanden", "bekleding_pcd"),
            ]:
                pcd = nearest[pcd_key]
                tanden = nearest[tanden_key]
                if pcd is None or pcd <= 0:
                    continue
                key = (band["naam"], uitvoering, motortype["code"])
                if key in seen:
                    continue
                seen.add(key)
                compat.append(
                    {
                        "band_naam": band["naam"],
                        "uitvoering_code": uitvoering,
                        "leverancier": motortype["leverancier"],
                        "motortype_code": motortype["code"],
                        "tandenaantal": tanden,
                        "pcd_mm": round(float(pcd), 4),
                    }
                )
    return compat


def upsert_import(parsed: ParsedData, dry_run: bool) -> dict[str, Any]:
    summary = {
        "dry_run": dry_run,
        "leveranciers": len(parsed.leveranciers),
        "uitvoering_types": len(parsed.uitvoering_types),
        "Aansluiting_types": len(parsed.Aansluiting_types),
        "band_types": len(parsed.band_types),
        "motortypes": len(parsed.motortypes),
        "motor_specs": len(parsed.motor_specs),
        "compatibiliteit": len(parsed.compatibiliteit),
        "warnings": parsed.warnings[:40],
    }
    if dry_run:
        return summary

    with get_db() as conn:
        cursor = conn.cursor()

        leverancier_ids: dict[str, int] = {}
        for naam in parsed.leveranciers:
            cursor.execute("SELECT id FROM dbo.tbl_leveranciers WHERE LOWER(naam) = LOWER(?)", naam)
            row = cursor.fetchone()
            if row:
                leverancier_ids[naam] = int(row[0])
                cursor.execute(
                    "UPDATE dbo.tbl_leveranciers SET actief = 1, updated_at = GETDATE() WHERE id = ?",
                    leverancier_ids[naam],
                )
            else:
                cursor.execute(
                    "INSERT INTO dbo.tbl_leveranciers (naam, actief, created_at) VALUES (?, 1, GETDATE())",
                    naam,
                )
                cursor.execute("SELECT id FROM dbo.tbl_leveranciers WHERE LOWER(naam) = LOWER(?)", naam)
                leverancier_ids[naam] = int(cursor.fetchone()[0])

        uitvoering_ids: dict[str, int] = {}
        for row in parsed.uitvoering_types:
            cursor.execute("SELECT id FROM dbo.tbl_uitvoering_types WHERE LOWER(code) = LOWER(?)", row["code"])
            existing = cursor.fetchone()
            if existing:
                uitvoering_ids[row["code"]] = int(existing[0])
                cursor.execute(
                    "UPDATE dbo.tbl_uitvoering_types SET naam = ?, actief = 1, updated_at = GETDATE() WHERE id = ?",
                    row["naam"],
                    uitvoering_ids[row["code"]],
                )
            else:
                cursor.execute(
                    "INSERT INTO dbo.tbl_uitvoering_types (naam, code, actief, created_at) VALUES (?, ?, 1, GETDATE())",
                    row["naam"],
                    row["code"],
                )
                cursor.execute("SELECT id FROM dbo.tbl_uitvoering_types WHERE LOWER(code) = LOWER(?)", row["code"])
                uitvoering_ids[row["code"]] = int(cursor.fetchone()[0])

        for row in parsed.Aansluiting_types:
            cursor.execute("SELECT id FROM dbo.tbl_Aansluiting_types WHERE LOWER(code) = LOWER(?)", row["code"])
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    "UPDATE dbo.tbl_Aansluiting_types SET naam = ?, actief = 1, updated_at = GETDATE() WHERE id = ?",
                    row["naam"],
                    int(existing[0]),
                )
            else:
                cursor.execute(
                    "INSERT INTO dbo.tbl_Aansluiting_types (naam, code, actief, created_at) VALUES (?, ?, 1, GETDATE())",
                    row["naam"],
                    row["code"],
                )

        band_ids: dict[str, int] = {}
        for row in parsed.band_types:
            cursor.execute("SELECT id FROM dbo.tbl_band_types WHERE LOWER(naam) = LOWER(?)", row["naam"])
            existing = cursor.fetchone()
            if existing:
                band_id = int(existing[0])
                band_ids[row["naam"]] = band_id
                cursor.execute(
                    """
                    UPDATE dbo.tbl_band_types
                    SET steek_mm = ?, dikte_tand_mm = ?, dikte_band_mm = ?,
                        wrijvingscoeff_rvs_droog = ?, wrijvingscoeff_rvs_nat = ?,
                        actief = 1, updated_at = GETDATE()
                    WHERE id = ?
                    """,
                    row["steek_mm"],
                    row["dikte_tand_mm"],
                    row["dikte_band_mm"],
                    row["wrijvingscoeff_rvs_droog"],
                    row["wrijvingscoeff_rvs_nat"],
                    band_id,
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO dbo.tbl_band_types (
                      naam, steek_mm, dikte_tand_mm, dikte_band_mm,
                      wrijvingscoeff_rvs_droog, wrijvingscoeff_rvs_nat, actief, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, 1, GETDATE())
                    """,
                    row["naam"],
                    row["steek_mm"],
                    row["dikte_tand_mm"],
                    row["dikte_band_mm"],
                    row["wrijvingscoeff_rvs_droog"],
                    row["wrijvingscoeff_rvs_nat"],
                )
                cursor.execute("SELECT id FROM dbo.tbl_band_types WHERE LOWER(naam) = LOWER(?)", row["naam"])
                band_ids[row["naam"]] = int(cursor.fetchone()[0])

        motortype_ids: dict[tuple[str, str], int] = {}
        for row in parsed.motortypes:
            leverancier_id = leverancier_ids[row["leverancier"]]
            cursor.execute(
                """
                SELECT id FROM dbo.tbl_motortypes
                WHERE leverancier_id = ? AND LOWER(code) = LOWER(?)
                """,
                leverancier_id,
                row["code"],
            )
            existing = cursor.fetchone()
            if existing:
                motortype_id = int(existing[0])
                motortype_ids[(row["leverancier"], row["code"])] = motortype_id
                cursor.execute(
                    """
                    UPDATE dbo.tbl_motortypes
                    SET diameter_nominaal_mm = ?, lengte_min_mm = ?, lengte_max_mm = ?,
                        actief = 1, updated_at = GETDATE()
                    WHERE id = ?
                    """,
                    row["diameter_nominaal_mm"],
                    row["lengte_min_mm"],
                    row["lengte_max_mm"],
                    motortype_id,
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO dbo.tbl_motortypes (
                        leverancier_id, code, diameter_nominaal_mm, lengte_min_mm, lengte_max_mm, actief, created_at
                    ) VALUES (?, ?, ?, ?, ?, 1, GETDATE())
                    """,
                    leverancier_id,
                    row["code"],
                    row["diameter_nominaal_mm"],
                    row["lengte_min_mm"],
                    row["lengte_max_mm"],
                )
                cursor.execute(
                    "SELECT id FROM dbo.tbl_motortypes WHERE leverancier_id = ? AND LOWER(code) = LOWER(?)",
                    leverancier_id,
                    row["code"],
                )
                motortype_ids[(row["leverancier"], row["code"])] = int(cursor.fetchone()[0])

        for row in parsed.motor_specs:
            motortype_id = motortype_ids[(row["leverancier"], row["motortype_code"])]
            cursor.execute(
                """
                SELECT id
                FROM dbo.tbl_motor_specs
                WHERE motortype_id = ?
                  AND ABS(vermogen_w - ?) < 0.0001
                  AND ABS(snelheid_ms - ?) < 0.0001
                  AND ISNULL(polen, -1) = ISNULL(?, -1)
                  AND LOWER(ISNULL(olie_type, '')) = LOWER(ISNULL(?, ''))
                """,
                motortype_id,
                row["vermogen_w"],
                row["snelheid_ms"],
                row["polen"],
                row["olie_type"],
            )
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    """
                    UPDATE dbo.tbl_motor_specs
                    SET force_n = ?, torque_nm = ?, actief = 1, updated_at = GETDATE()
                    WHERE id = ?
                    """,
                    row["force_n"],
                    row["torque_nm"],
                    int(existing[0]),
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO dbo.tbl_motor_specs (
                      motortype_id, vermogen_w, snelheid_ms, polen, force_n, torque_nm, olie_type, actief, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, GETDATE())
                    """,
                    motortype_id,
                    row["vermogen_w"],
                    row["snelheid_ms"],
                    row["polen"],
                    row["force_n"],
                    row["torque_nm"],
                    row["olie_type"],
                )

        for row in parsed.compatibiliteit:
            band_id = band_ids[row["band_naam"]]
            uitvoering_id = uitvoering_ids[row["uitvoering_code"]]
            motortype_id = motortype_ids[(row["leverancier"], row["motortype_code"])]
            cursor.execute(
                """
                SELECT id FROM dbo.tbl_band_motor_compatibiliteit
                WHERE band_type_id = ? AND uitvoering_type_id = ? AND motortype_id = ?
                """,
                band_id,
                uitvoering_id,
                motortype_id,
            )
            existing = cursor.fetchone()
            if existing:
                cursor.execute(
                    """
                    UPDATE dbo.tbl_band_motor_compatibiliteit
                    SET tandenaantal = ?, pcd_mm = ?, actief = 1, updated_at = GETDATE()
                    WHERE id = ?
                    """,
                    row["tandenaantal"],
                    row["pcd_mm"],
                    int(existing[0]),
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO dbo.tbl_band_motor_compatibiliteit (
                      band_type_id, uitvoering_type_id, motortype_id, tandenaantal, pcd_mm, actief, created_at
                    ) VALUES (?, ?, ?, ?, ?, 1, GETDATE())
                    """,
                    band_id,
                    uitvoering_id,
                    motortype_id,
                    row["tandenaantal"],
                    row["pcd_mm"],
                )

        conn.commit()
    return summary


def run_import(workbook_path: Path, dry_run: bool) -> dict[str, Any]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    parsed = parse_workbook(workbook_path)
    summary = upsert_import(parsed, dry_run=dry_run)
    summary["workbook"] = str(workbook_path)
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Ammdrive workbook into SQL catalog tables.")
    parser.add_argument("--file", dest="workbook", default=str(DEFAULT_XLSX_PATH))
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true")
    mode.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    summary = run_import(workbook_path, dry_run=bool(args.dry_run))
    print(json.dumps(summary, ensure_ascii=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
